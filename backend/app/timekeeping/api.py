from __future__ import annotations

from collections import defaultdict
from calendar import monthrange

from datetime import date, datetime, timezone
from typing import Dict, List, Optional
from .schemas import CrewSegmentCreate, CrewSegmentClose, CrewSegmentOut, SegmentStartIn, DailyReportOut, DailyEmployeeTotal, DailySiteTotal, DailyCrewLogTotal, RangeReportOut, RangeDayTotal, RangeVehicleTotal, RangeEmployeeTotal, RangeSiteTotal

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import bindparam
from sqlalchemy import and_
from pydantic import BaseModel, Field


class DayCrewLogOut(BaseModel):
    crew_log_id: int
    work_date: date
    site_id: Optional[int] = None
    site_name: Optional[str] = None
    vehicle_id: Optional[int] = None
    vehicle_plate: Optional[str] = None
    employees: List[str] = []
    work_minutes: int = 0
    work_hours: float = 0.0
    travel_minutes: int = 0
    travel_hours: float = 0.0
    km: float = 0.0
    segments_count: int = 0

class DayReportOut(BaseModel):
    date: date
    crew_logs: List[DayCrewLogOut] = []

class EmployeeDayOut(BaseModel):
    date: date
    work_minutes: int = 0
    work_hours: float = 0.0
    travel_minutes: int = 0
    travel_hours: float = 0.0
    km: float = 0.0

class EmployeeReportOut(BaseModel):
    employee_id: int
    employee_name: str
    date_from: date
    date_to: date
    total_work_hours: float = 0.0
    total_travel_hours: float = 0.0
    total_km: float = 0.0
    days: List[EmployeeDayOut] = []
from sqlalchemy.orm import Session

from app.db import SessionLocal
from app.timekeeping.models import (
    TkCrewLog,
    TkCrewLogMember,
    TkCrewWorkSegment,
    TkEmployee,
    TkVehicle,
    TkSite,
    TkSegmentType,
)

router = APIRouter(prefix="/timekeeping", tags=["timekeeping"])


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# -----------------------
# Employees
# -----------------------

class EmployeeCreate(BaseModel):
    full_name: str = Field(min_length=2, max_length=200)
    user_id: Optional[int] = None


class EmployeeOut(BaseModel):
    id: int
    full_name: str
    user_id: Optional[int]
    is_active: bool

    class Config:
        from_attributes = True


@router.post("/employees", response_model=EmployeeOut)
def create_employee(payload: EmployeeCreate, db: Session = Depends(get_db)):
    emp = TkEmployee(full_name=payload.full_name, user_id=payload.user_id, is_active=True)
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


@router.get("/employees", response_model=List[EmployeeOut])
def list_employees(db: Session = Depends(get_db)):
    return (
        db.query(TkEmployee)
        .filter(TkEmployee.is_active == True)  # noqa: E712
        .order_by(TkEmployee.full_name.asc())
        .all()
    )


# -----------------------
# Vehicles
# -----------------------

class VehicleCreate(BaseModel):
    plate: str = Field(min_length=2, max_length=32)
    make_model: Optional[str] = None
    navisoft_device_id: Optional[str] = None


class VehicleOut(BaseModel):
    id: int
    plate: str
    make_model: Optional[str]
    navisoft_device_id: Optional[str]
    is_active: bool

    class Config:
        from_attributes = True


@router.post("/vehicles", response_model=VehicleOut)
def create_vehicle(payload: VehicleCreate, db: Session = Depends(get_db)):
    v = TkVehicle(
        plate=payload.plate,
        make_model=payload.make_model,
        navisoft_device_id=payload.navisoft_device_id,
        is_active=True,
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    return v


@router.get("/vehicles", response_model=List[VehicleOut])
def list_vehicles(db: Session = Depends(get_db)):
    return (
        db.query(TkVehicle)
        .filter(TkVehicle.is_active == True)  # noqa: E712
        .order_by(TkVehicle.plate.asc())
        .all()
    )


# -----------------------
# Sites
# -----------------------

class SiteAdHocCreate(BaseModel):
    name: str = Field(min_length=2, max_length=200)
    lat: float
    lng: float
    radius_m: int = 500


class SiteOut(BaseModel):
    id: int
    name: str
    lat: Optional[float]
    lng: Optional[float]
    radius_m: Optional[int]
    is_ad_hoc: bool

    class Config:
        from_attributes = True


@router.post("/sites/ad-hoc", response_model=SiteOut)
def create_site_ad_hoc(payload: SiteAdHocCreate, db: Session = Depends(get_db)):
    s = TkSite(
        name=payload.name,
        lat=payload.lat,
        lng=payload.lng,
        radius_m=payload.radius_m,
        is_ad_hoc=True,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


@router.get("/sites", response_model=List[SiteOut])
def list_sites(db: Session = Depends(get_db)):
    return db.query(TkSite).order_by(TkSite.name.asc()).all()


# -----------------------
# Crew Logs
# -----------------------

class CrewLogCreate(BaseModel):
    work_date: date
    vehicle_id: int
    created_by_employee_id: Optional[int] = None


class CrewLogOut(BaseModel):
    id: int
    work_date: date
    vehicle_id: int
    created_by_employee_id: Optional[int] = None

    class Config:
        from_attributes = True


@router.post("/crew-logs", response_model=CrewLogOut)
def create_crew_log(payload: CrewLogCreate, db: Session = Depends(get_db)):
    # ochrona przed duplikatem (cz?sty pow?d 500, je?li jest UNIQUE)
    existing = (
        db.query(TkCrewLog)
        .filter(TkCrewLog.work_date == payload.work_date)
        .filter(TkCrewLog.vehicle_id == payload.vehicle_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail=f"Crew log already exists (id={existing.id}).")

    log = TkCrewLog(
        work_date=payload.work_date,
        vehicle_id=payload.vehicle_id,
        created_by_employee_id=payload.created_by_employee_id,
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    return log


@router.get("/crew-logs", response_model=List[CrewLogOut])
def list_crew_logs(
    work_date: Optional[date] = None,
    vehicle_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(TkCrewLog)
    if work_date is not None:
        q = q.filter(TkCrewLog.work_date == work_date)
    if vehicle_id is not None:
        q = q.filter(TkCrewLog.vehicle_id == vehicle_id)
    return q.order_by(TkCrewLog.id.desc()).all()


class CrewMemberCreate(BaseModel):
    employee_id: int


class CrewMemberOut(BaseModel):
    id: int
    crew_log_id: int
    employee_id: int

    class Config:
        from_attributes = True


@router.get("/crew-logs/{log_id}/members", response_model=List[CrewMemberOut])
def list_members(log_id: int, db: Session = Depends(get_db)):
    return (
        db.query(TkCrewLogMember)
        .filter(TkCrewLogMember.crew_log_id == log_id)
        .order_by(TkCrewLogMember.id.asc())
        .all()
    )


@router.post("/crew-logs/{log_id}/members", response_model=CrewMemberOut)
def add_member(log_id: int, payload: CrewMemberCreate, db: Session = Depends(get_db)):
    log = db.query(TkCrewLog).filter(TkCrewLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Crew log not found")

    emp = db.query(TkEmployee).filter(TkEmployee.id == payload.employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # blokada duplikatu cz?onka
    exists = (
        db.query(TkCrewLogMember)
        .filter(TkCrewLogMember.crew_log_id == log_id)
        .filter(TkCrewLogMember.employee_id == payload.employee_id)
        .first()
    )
    if exists:
        return exists

    m = TkCrewLogMember(crew_log_id=log_id, employee_id=payload.employee_id)
    db.add(m)
    db.commit()
    db.refresh(m)
    return m


# -----------------------
# Segments
# -----------------------
@router.get("/crew-logs/{log_id}/segments", response_model=List[CrewSegmentOut])
def list_segments(log_id: int, db: Session = Depends(get_db)):
    return (
        db.query(TkCrewWorkSegment)
        .filter(TkCrewWorkSegment.crew_log_id == log_id)
        .order_by(TkCrewWorkSegment.id.asc())
        .all()
    )


@router.post("/crew-logs/{log_id}/segments", response_model=CrewSegmentOut)
def add_segment(log_id: int, payload: CrewSegmentCreate, db: Session = Depends(get_db)):
    log = db.query(TkCrewLog).filter(TkCrewLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Crew log not found")

    open_seg = (
        db.query(TkCrewWorkSegment)
        .filter(TkCrewWorkSegment.crew_log_id == log_id)
        .filter(TkCrewWorkSegment.end_at.is_(None))
        .first()
    )
    if open_seg:
        raise HTTPException(
            status_code=409,
            detail=f"Open segment already exists (segment_id={open_seg.id}). Close it first.",
        )

    site = db.query(TkSite).filter(TkSite.id == payload.site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    if site.lat is None or site.lng is None:
        raise HTTPException(status_code=422, detail="Site is missing lat/lng")

    
    AUTO_TRAVEL_GAP = True
    
    try:
        is_work = (getattr(payload, "segment_type", None) in (None, "work", TkSegmentType.work))
    except Exception:
        is_work = True
    
    if is_work and payload.start_at:
        last = (
            db.query(TkCrewWorkSegment)
            .filter(TkCrewWorkSegment.crew_log_id == log_id)
            .filter(TkCrewWorkSegment.end_at.isnot(None))
            .order_by(TkCrewWorkSegment.end_at.desc())
            .first()
        )
        last_end = last.end_at if last else None
        start_dt = payload.start_at
        try:
            if last_end is not None and getattr(last_end, "tzinfo", None) is not None:
                last_end = last_end.replace(tzinfo=None)
            if start_dt is not None and getattr(start_dt, "tzinfo", None) is not None:
                start_dt = start_dt.replace(tzinfo=None)
        except Exception:
            pass
        if last_end and start_dt and last_end < start_dt:
            gap_min = int((start_dt - last_end).total_seconds() // 60)
            if gap_min > 0:
                travel = TkCrewWorkSegment(
                    crew_log_id=log_id,
                    site_id=payload.site_id,
                    segment_type=TkSegmentType.travel,
                    start_at=last_end,
                    end_at=start_dt,
                    start_lat=site.lat,
                    start_lng=site.lng,
                    end_lat=site.lat,
                    end_lng=site.lng,
                )
                db.add(travel)
    
    seg = TkCrewWorkSegment(
        crew_log_id=log_id,
        site_id=payload.site_id,
        segment_type=getattr(payload, "segment_type", None) or TkSegmentType.work,
        start_at=payload.start_at,
        end_at=payload.end_at,
        start_lat=site.lat,
        start_lng=site.lng,
        end_lat=site.lat,
        end_lng=site.lng,
    )
    db.add(seg)
    db.commit()
    db.refresh(seg)
    return seg

@router.post("/crew-logs/{log_id}/segments/start", response_model=CrewSegmentOut)
def start_segment(log_id: int, payload: SegmentStartIn, db: Session = Depends(get_db)):
    now = datetime.utcnow()
    return add_segment(
        log_id=log_id,
        payload=CrewSegmentCreate(site_id=payload.site_id, start_at=now, end_at=None),
        db=db,
    )


@router.patch("/crew-logs/{log_id}/segments/{segment_id}/close", response_model=CrewSegmentOut)
def close_segment(log_id: int, segment_id: int, payload: CrewSegmentClose, db: Session = Depends(get_db)):
    seg = (
        db.query(TkCrewWorkSegment)
        .filter(TkCrewWorkSegment.id == segment_id)
        .filter(TkCrewWorkSegment.crew_log_id == log_id)
        .first()
    )
    if not seg:
        raise HTTPException(status_code=404, detail="Segment not found")

    if seg.end_at is not None:
        raise HTTPException(status_code=409, detail="Segment already closed")

    if seg.site_id is None:
        raise HTTPException(status_code=422, detail="Segment is missing site_id")

    site = db.query(TkSite).filter(TkSite.id == seg.site_id).first()
    if not site or site.lat is None or site.lng is None:
        raise HTTPException(status_code=422, detail="Site is missing lat/lng")

    if seg.start_lat is None or seg.start_lng is None:
        seg.start_lat = site.lat
        seg.start_lng = site.lng

    seg.end_at = payload.end_at or datetime.now(timezone.utc)
    seg.end_lat = site.lat
    seg.end_lng = site.lng

    dist = float(getattr(payload, "distance_km", 0.0) or 0.0)
    if str(getattr(seg, "segment_type", None) or "").lower() == "travel":
        seg.distance_km = dist
    else:
        seg.distance_km = 0.0
    db.add(seg)
    db.commit()
    db.refresh(seg)
    return seg

@router.patch("/crew-logs/{log_id}/segments/stop", response_model=CrewSegmentOut)
def stop_segment(log_id: int, db: Session = Depends(get_db)):
    seg = (
        db.query(TkCrewWorkSegment)
        .filter(TkCrewWorkSegment.crew_log_id == log_id)
        .filter(TkCrewWorkSegment.end_at.is_(None))
        .order_by(TkCrewWorkSegment.id.desc())
        .first()
    )
    if not seg:
        raise HTTPException(status_code=404, detail="No open segment found")

    return close_segment(
        log_id=log_id,
        segment_id=seg.id,
        payload=CrewSegmentClose(end_at=datetime.now(timezone.utc)),
        db=db,
    )


# -----------------------
# Summary
# -----------------------

class CrewLogSummaryOut(BaseModel):
    crew_log_id: int
    work_date: date
    vehicle_id: int
    total_minutes: int
    by_site_minutes: Dict[int, int]
    by_employee_minutes: Dict[int, int]


@router.get("/crew-logs/{log_id}/summary", response_model=CrewLogSummaryOut)
def crew_log_summary(log_id: int, db: Session = Depends(get_db)):
    log = db.query(TkCrewLog).filter(TkCrewLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Crew log not found")

    segs = (
        db.query(TkCrewWorkSegment)
        .filter(TkCrewWorkSegment.crew_log_id == log_id)
        .filter(TkCrewWorkSegment.end_at.isnot(None))
        .order_by(TkCrewWorkSegment.start_at.asc())
        .all()
    )

    by_site = defaultdict(int)
    total = 0

    for seg in segs:
        minutes = int((seg.end_at - seg.start_at).total_seconds() // 60)
        is_travel = (getattr(seg, "segment_type", None) == TkSegmentType.travel)
        if minutes < 0:
            minutes = 0
        total += minutes
        by_site[int(seg.site_id)] += minutes

    members = (
        db.query(TkCrewLogMember)
        .filter(TkCrewLogMember.crew_log_id == log_id)
        .all()
    )

    # MVP: ka?demu cz?onkowi przypisujemy total (to p??niej rozbijemy proporcjonalnie)
    by_emp = {int(m.employee_id): int(total) for m in members}

    return CrewLogSummaryOut(
        crew_log_id=log.id,
        work_date=log.work_date,
        vehicle_id=log.vehicle_id,
        total_minutes=int(total),
        by_site_minutes=dict(by_site),
        by_employee_minutes=by_emp,
    )












# -----------------------
# Reports
# -----------------------

@router.get("/reports/daily", response_model=DailyReportOut)
def report_daily(
    work_date: date,
    vehicle_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = (
        db.query(
            TkCrewWorkSegment.crew_log_id,
            TkCrewWorkSegment.site_id,
            TkCrewWorkSegment.start_at,
            TkCrewWorkSegment.end_at,
            TkCrewLog.vehicle_id.label("vehicle_id"),
        )
        .join(TkCrewLog, TkCrewLog.id == TkCrewWorkSegment.crew_log_id)
        .filter(TkCrewLog.work_date == work_date)
        .filter(TkCrewWorkSegment.end_at.isnot(None))
    )

    if vehicle_id is not None:
        q = q.filter(TkCrewLog.vehicle_id == vehicle_id)

    segments = q.all()

    crew_log_ids = sorted({s.crew_log_id for s in segments})
    site_ids = sorted({s.site_id for s in segments})

    sites_by_id = {}
    if site_ids:
        for s in db.query(TkSite).filter(TkSite.id.in_(site_ids)).all():
            sites_by_id[s.id] = s.name

    crewlog_vehicle = {}
    if crew_log_ids:
        for cl in db.query(TkCrewLog).filter(TkCrewLog.id.in_(crew_log_ids)).all():
            crewlog_vehicle[cl.id] = cl.vehicle_id

    members_by_log = defaultdict(list)
    if crew_log_ids:
        members = (
            db.query(TkCrewLogMember)
            .filter(TkCrewLogMember.crew_log_id.in_(crew_log_ids))
            .all()
        )
        for m in members:
            members_by_log[m.crew_log_id].append(m.employee_id)

    employees_by_id = {}
    if crew_log_ids:
        emp_ids = sorted({eid for ids in members_by_log.values() for eid in ids})
        if emp_ids:
            for e in db.query(TkEmployee).filter(TkEmployee.id.in_(emp_ids)).all():
                employees_by_id[e.id] = e.full_name

    site_totals: Dict[int, Dict[str, int]] = defaultdict(lambda: {"minutes": 0, "segments": 0, "work_minutes": 0, "travel_minutes": 0})
    crewlog_totals: Dict[int, Dict[str, int]] = defaultdict(lambda: {"minutes": 0, "segments": 0, "work_minutes": 0, "travel_minutes": 0})
    employee_totals: Dict[int, Dict[str, int]] = defaultdict(lambda: {"minutes": 0, "segments": 0, "work_minutes": 0, "travel_minutes": 0})

    total_minutes = 0

    for s in segments:
        start_at = s.start_at
        end_at = s.end_at
        if not end_at:
            continue

        minutes = int(max(0, (end_at - start_at).total_seconds() // 60))
        total_minutes += minutes

        site_totals[s.site_id]["minutes"] += minutes
        is_travel = (getattr(seg, "segment_type", None) == TkSegmentType.travel)
        if is_travel:
            site_totals[s.site_id]["travel_minutes"] += minutes
        else:
            site_totals[s.site_id]["work_minutes"] += minutes
        site_totals[s.site_id]["segments"] += 1

        crewlog_totals[s.crew_log_id]["minutes"] += minutes

        if is_travel:
            crewlog_totals[s.crew_log_id]["travel_minutes"] += minutes
        else:
            crewlog_totals[s.crew_log_id]["work_minutes"] += minutes
        if is_travel:
            crewlog_totals[s.crew_log_id]["travel_minutes"] += minutes
        else:
            crewlog_totals[s.crew_log_id]["work_minutes"] += minutes
        crewlog_totals[s.crew_log_id]["segments"] += 1

        emp_ids = members_by_log.get(s.crew_log_id, [])
        if employee_id is not None:
            emp_ids = [x for x in emp_ids if x == employee_id]

        if emp_ids:
            per_emp = minutes // len(emp_ids) if len(emp_ids) > 0 else 0
            for eid in emp_ids:
                employee_totals[eid]["minutes"] += per_emp
                is_travel = (getattr(seg, "segment_type", None) == TkSegmentType.travel)
                if is_travel:
                    employee_totals[eid]["travel_minutes"] += per_emp
                else:
                    employee_totals[eid]["work_minutes"] += per_emp
                employee_totals[eid]["segments"] += 1

    employees_out = [
        DailyEmployeeTotal(
            employee_id=eid,
            full_name=employees_by_id.get(eid, f"Employee {eid}"),
            minutes=vals["minutes"],
            work_minutes=vals.get("work_minutes", 0),
            travel_minutes=vals.get("travel_minutes", 0),
            segments=vals["segments"],
        )
        for eid, vals in sorted(employee_totals.items(), key=lambda kv: (-(kv[1]["minutes"]), kv[0]))
    ]

    sites_out = [
        DailySiteTotal(
            site_id=sid,
            name=sites_by_id.get(sid, f"Site {sid}"),
            minutes=vals["minutes"],
            work_minutes=vals.get("work_minutes", 0),
            travel_minutes=vals.get("travel_minutes", 0),
            segments=vals["segments"],
        )
        for sid, vals in sorted(site_totals.items(), key=lambda kv: (-(kv[1]["minutes"]), kv[0]))
    ]

    crew_logs_out = [
        DailyCrewLogTotal(
            crew_log_id=clid,
            vehicle_id=crewlog_vehicle.get(clid, 0),
            minutes=vals["minutes"],
            work_minutes=vals.get("work_minutes", 0),
            travel_minutes=vals.get("travel_minutes", 0),
            segments=vals["segments"],
        )
        for clid, vals in sorted(crewlog_totals.items(), key=lambda kv: (-(kv[1]["minutes"]), kv[0]))
    ]

    def _val(x, k):
        return x.get(k, 0) if isinstance(x, dict) else getattr(x, k, 0)

    minutes_total = sum(_val(x, "minutes") for x in crew_logs_out) if crew_logs_out else 0
    work_minutes_total = sum(_val(x, "work_minutes") for x in crew_logs_out) if crew_logs_out else 0
    travel_minutes_total = sum(_val(x, "travel_minutes") for x in crew_logs_out) if crew_logs_out else 0

    return DailyReportOut(
        work_date=work_date,
        total_minutes=minutes_total,
        work_minutes=work_minutes_total,
        travel_minutes=travel_minutes_total,
        employees=employees_out,
        sites=sites_out,
        crew_logs=crew_logs_out,
        minutes=minutes,

    )
@router.get("/reports/range", response_model=RangeReportOut)
def report_range(
date_from: date,
date_to: date,
vehicle_id: Optional[int] = None,
employee_id: Optional[int] = None,
db: Session = Depends(get_db),
):
    res = _aggregate_range(db, date_from, date_to, vehicle_id=vehicle_id, employee_id=employee_id)
    if res is None:
        return {
            "total_minutes": 0,
        "date_from": str(date_from),
        "date_to": str(date_to),
            "days": [],
            "employees": [],
            "sites": [],
            "vehicles": [],
        }
    return res
@router.get("/reports/weekly", response_model=RangeReportOut)
def report_weekly(
    week_start: date,
    vehicle_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    date_from = week_start
    date_to = date(week_start.year, week_start.month, week_start.day)  # defensive
    date_to = date_from.fromordinal(date_from.toordinal() + 6)
    return _aggregate_range(db, date_from, date_to, vehicle_id=vehicle_id, employee_id=employee_id)
@router.get("/reports/monthly", response_model=RangeReportOut)
def report_monthly(
    year: int,
    month: int,
    vehicle_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month must be 1..12")
    last_day = monthrange(year, month)[1]
    date_from = date(year, month, 1)
    date_to = date(year, month, last_day)
    return _aggregate_range(db, date_from, date_to, vehicle_id=vehicle_id, employee_id=employee_id)
# -----------------------
# CSV Exports
# -----------------------

from io import StringIO
import csv
from fastapi.responses import Response

def _csv_response(filename: str, header: list[str], rows: list[list[object]]) -> Response:
    buf = StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(header)
    for r in rows:
        w.writerow(r)

    content = "\ufeff" + buf.getvalue()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _date_add_days(d: date, days: int) -> date:
    return date.fromordinal(d.toordinal() + days)

def _range_report(db, date_from, date_to, vehicle_id=None, employee_id=None):
    return _aggregate_range(db, date_from, date_to, vehicle_id=vehicle_id, employee_id=employee_id)

def _aggregate_range(db, date_from, date_to, vehicle_id=None, employee_id=None):
    from datetime import date as _date
    from sqlalchemy import text
    from app.timekeeping.time_utils import ceil_minutes_to_quarters, split_work_travel_hours

    def _d(x):
        if isinstance(x, _date):
            return x.isoformat()
        return str(x)

    def _round15_minutes(minutes):
        q = ceil_minutes_to_quarters(minutes)
        return int(q) * 15

    date_from_s = _d(date_from)
    date_to_s = _d(date_to)

    base_where = "l.work_date >= :date_from AND l.work_date <= :date_to AND ws.start_at IS NOT NULL AND ws.end_at IS NOT NULL"
    params = {"date_from": date_from_s, "date_to": date_to_s}

    if vehicle_id is not None:
        base_where += " AND l.vehicle_id = :vehicle_id"
        params["vehicle_id"] = int(vehicle_id)

    emp_filter_join = ""
    emp_filter_where = ""
    if employee_id is not None:
        emp_filter_join = "JOIN tk_crew_log_members lm_f ON lm_f.crew_log_id = l.id"
        emp_filter_where = " AND lm_f.employee_id = :employee_id"
        params["employee_id"] = int(employee_id)

    minutes_expr = "(julianday(ws.end_at) - julianday(ws.start_at)) * 1440.0"
    work_min_expr = f"SUM(CASE WHEN ws.segment_type = 'work' THEN {minutes_expr} ELSE 0.0 END)"
    travel_min_expr = f"SUM(CASE WHEN ws.segment_type = 'travel' THEN {minutes_expr} ELSE 0.0 END)"
    total_min_expr = f"SUM({minutes_expr})"

    total_sql = text(f"""
SELECT COALESCE({total_min_expr}, 0.0) AS minutes,
       COALESCE({work_min_expr}, 0.0) AS work_minutes,
       COALESCE({travel_min_expr}, 0.0) AS travel_minutes
FROM tk_crew_work_segments ws
JOIN tk_crew_logs l ON l.id = ws.crew_log_id
{emp_filter_join}
WHERE {base_where}{emp_filter_where}
""")

    days_sql = text(f"""
SELECT l.work_date AS work_date,
       COALESCE({total_min_expr}, 0.0) AS minutes,
       COALESCE({work_min_expr}, 0.0) AS work_minutes,
       COALESCE({travel_min_expr}, 0.0) AS travel_minutes,
       COUNT(ws.id) AS segments
FROM tk_crew_work_segments ws
JOIN tk_crew_logs l ON l.id = ws.crew_log_id
{emp_filter_join}
WHERE {base_where}{emp_filter_where}
GROUP BY l.work_date
ORDER BY l.work_date
""")

    sites_sql = text(f"""
SELECT ws.site_id AS site_id,
       COALESCE(s.name, '') AS name,
       COALESCE({total_min_expr}, 0.0) AS minutes,
       COALESCE({work_min_expr}, 0.0) AS work_minutes,
       COALESCE({travel_min_expr}, 0.0) AS travel_minutes,
       COUNT(ws.id) AS segments
FROM tk_crew_work_segments ws
JOIN tk_crew_logs l ON l.id = ws.crew_log_id
LEFT JOIN tk_sites s ON s.id = ws.site_id
{emp_filter_join}
WHERE {base_where}{emp_filter_where}
GROUP BY ws.site_id, s.name
ORDER BY minutes DESC
""")

    vehicles_sql = text(f"""
SELECT l.vehicle_id AS vehicle_id,
       COALESCE(v.plate, '') AS plate,
       COALESCE(SUM(ws.distance_km), 0.0) AS km,
       COALESCE({total_min_expr}, 0.0) AS minutes,
       COALESCE({work_min_expr}, 0.0) AS work_minutes,
       COALESCE({travel_min_expr}, 0.0) AS travel_minutes,
       COUNT(ws.id) AS segments
FROM tk_crew_work_segments ws
JOIN tk_crew_logs l ON l.id = ws.crew_log_id
LEFT JOIN tk_vehicles v ON v.id = l.vehicle_id
{emp_filter_join}
WHERE {base_where}{emp_filter_where}
GROUP BY l.vehicle_id, v.plate
ORDER BY km DESC
""")

    employees_sql = text(f"""
SELECT e.id AS employee_id,
       e.full_name AS full_name,
       COALESCE({total_min_expr}, 0.0) AS minutes,
       COALESCE({work_min_expr}, 0.0) AS work_minutes,
       COALESCE({travel_min_expr}, 0.0) AS travel_minutes,
       COUNT(ws.id) AS segments
FROM tk_crew_work_segments ws
JOIN tk_crew_logs l ON l.id = ws.crew_log_id
JOIN tk_crew_log_members lm ON lm.crew_log_id = l.id
JOIN tk_employees e ON e.id = lm.employee_id
{emp_filter_join}
WHERE {base_where}{emp_filter_where}
GROUP BY e.id, e.full_name
ORDER BY minutes DESC
""")

    total_row = db.execute(total_sql, params).mappings().first()
    raw_total_min = total_row["minutes"] if total_row else 0.0
    raw_work_min = total_row["work_minutes"] if total_row else 0.0
    raw_travel_min = total_row["travel_minutes"] if total_row else 0.0

    total_minutes = _round15_minutes(raw_total_min)
    work_minutes = _round15_minutes(raw_work_min)
    travel_minutes = _round15_minutes(raw_travel_min)

    total_hours, total_work_hours, total_travel_hours = split_work_travel_hours(raw_work_min, raw_travel_min)

    days_rows = db.execute(days_sql, params).mappings().all()
    sites_rows = db.execute(sites_sql, params).mappings().all()
    vehicles_rows = db.execute(vehicles_sql, params).mappings().all()
    employees_rows = db.execute(employees_sql, params).mappings().all()

    days = []
    for r in days_rows:
        rm = float(r["minutes"] or 0.0)
        rwm = float(r["work_minutes"] or 0.0)
        rtm = float(r["travel_minutes"] or 0.0)
        h, wh, th = split_work_travel_hours(rwm, rtm)
        days.append({
            "work_date": _d(r["work_date"]),
            "minutes": _round15_minutes(rm),
            "work_minutes": _round15_minutes(rwm),
            "travel_minutes": _round15_minutes(rtm),
            "hours": h,
            "work_hours": wh,
            "travel_hours": th,
            "segments": int(r["segments"] or 0),
        })

    sites = []
    for r in sites_rows:
        rm = float(r["minutes"] or 0.0)
        rwm = float(r["work_minutes"] or 0.0)
        rtm = float(r["travel_minutes"] or 0.0)
        h, wh, th = split_work_travel_hours(rwm, rtm)
        sites.append({
            "site_id": int(r["site_id"]) if r["site_id"] is not None else 0,
            "name": r["name"] or "",
            "minutes": _round15_minutes(rm),
            "work_minutes": _round15_minutes(rwm),
            "travel_minutes": _round15_minutes(rtm),
            "hours": h,
            "work_hours": wh,
            "travel_hours": th,
            "segments": int(r["segments"] or 0),
        })

    vehicles = []
    for r in vehicles_rows:
        rm = float(r["minutes"] or 0.0)
        rwm = float(r["work_minutes"] or 0.0)
        rtm = float(r["travel_minutes"] or 0.0)
        h, wh, th = split_work_travel_hours(rwm, rtm)
        vehicles.append({
            "vehicle_id": int(r["vehicle_id"]) if r["vehicle_id"] is not None else 0,
            "plate": r["plate"] or "",
            "km": float(r["km"] or 0.0),
            "minutes": _round15_minutes(rm),
            "work_minutes": _round15_minutes(rwm),
            "travel_minutes": _round15_minutes(rtm),
            "hours": h,
            "work_hours": wh,
            "travel_hours": th,
            "segments": int(r["segments"] or 0),
        })

    employees = []
    for r in employees_rows:
        rm = float(r["minutes"] or 0.0)
        rwm = float(r["work_minutes"] or 0.0)
        rtm = float(r["travel_minutes"] or 0.0)
        h, wh, th = split_work_travel_hours(rwm, rtm)
        employees.append({
            "employee_id": int(r["employee_id"]),
            "full_name": r["full_name"] or "",
            "minutes": _round15_minutes(rm),
            "work_minutes": _round15_minutes(rwm),
            "travel_minutes": _round15_minutes(rtm),
            "hours": h,
            "work_hours": wh,
            "travel_hours": th,
            "segments": int(r["segments"] or 0),
        })

    return {
        "date_from": date_from_s,
        "date_to": date_to_s,
        "total_minutes": int(total_minutes),
        "work_minutes": int(work_minutes),
        "travel_minutes": int(travel_minutes),
        "total_hours": float(total_hours),
        "work_hours": float(total_work_hours),
        "travel_hours": float(total_travel_hours),
        "days": days,
        "employees": employees,
        "sites": sites,
        "vehicles": vehicles,
    }









# -----------------------
# Report: Day (per crew log)
# -----------------------

@router.get("/reports/day", response_model=DayReportOut)
def report_day(
    date: date,
    db: Session = Depends(get_db),
):
    logs = (
        db.query(TkCrewLog)
        .filter(TkCrewLog.work_date == date)
        .all()
    )

    result = []

    for log in logs:
        segments = (
            db.query(TkCrewWorkSegment)
            .filter(TkCrewWorkSegment.crew_log_id == log.id)
            .filter(TkCrewWorkSegment.end_at.isnot(None))
            .all()
        )

        work_minutes = 0
        travel_minutes = 0
        km = 0.0

        for seg in segments:
            if not seg.start_at or not seg.end_at:
                continue

            minutes = int((seg.end_at - seg.start_at).total_seconds() // 60)
            if minutes < 0:
                minutes = 0

            minutes_15 = ((minutes + 14) // 15) * 15

            if str(getattr(seg, "segment_type", None) or "").lower() == "travel":
                travel_minutes += minutes_15
                km += float(seg.distance_km or 0.0)
            else:
                work_minutes += minutes_15

        members = (
            db.query(TkCrewLogMember)
            .filter(TkCrewLogMember.crew_log_id == log.id)
            .all()
        )

        employee_names = []
        if members:
            emps = (
                db.query(TkEmployee)
                .filter(TkEmployee.id.in_([m.employee_id for m in members]))
                .all()
            )
            employee_names = [e.full_name for e in emps]

        site_id = None
        site_name = None
        by_site_minutes = {}

        for seg in segments:
            sid = getattr(seg, "site_id", None)
            if sid is None:
                continue
            minutes = int((seg.end_at - seg.start_at).total_seconds() // 60) if seg.start_at and seg.end_at else 0
            if minutes < 0:
                minutes = 0
            minutes_15 = ((minutes + 14) // 15) * 15
            by_site_minutes[int(sid)] = by_site_minutes.get(int(sid), 0) + int(minutes_15)

        if by_site_minutes:
            site_id = max(by_site_minutes.items(), key=lambda kv: kv[1])[0]
            site = db.query(TkSite).filter(TkSite.id == site_id).first()
            if site:
                site_name = site.name

        vehicle_plate = None
        if log.vehicle_id:
            veh = db.query(TkVehicle).filter(TkVehicle.id == log.vehicle_id).first()
            if veh:
                vehicle_plate = veh.plate

        result.append({
            "crew_log_id": log.id,
            "work_date": log.work_date,
            "site_id": site_id,
            "site_name": site_name,
            "vehicle_id": log.vehicle_id,
            "vehicle_plate": vehicle_plate,
            "employees": employee_names,
            "work_minutes": work_minutes,
            "work_hours": round(work_minutes / 60, 2),
            "travel_minutes": travel_minutes,
            "travel_hours": round(travel_minutes / 60, 2),
            "km": round(km, 2),
            "segments_count": len(segments),
        })

    return {
        "date": date,
        "crew_logs": result,
    }






@router.get("/reports/day.xlsx")
def report_day_xlsx(
    date: date,
    rate_per_km: float = 0.0,
    db: Session = Depends(get_db),
):
    data = report_day(date=date, db=db)
    rep_date = str(data.get("date", date))
    crew_logs = data.get("crew_logs", [])

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Day"

    headers = [
        "crew_log_id",
        "work_date",
        "site_id",
        "site_name",
        "vehicle_id",
        "vehicle_plate",
        "employees",
        "work_hours",
        "travel_hours",
        "km",
        "segments_count",
        "travel_cost",
    ]
    ws.append(headers)

    total_work = 0.0
    total_travel = 0.0
    total_km = 0.0
    total_cost = 0.0

    for row in data.get("crew_logs", []):
        employees = ", ".join(row.get("employees") or [])
        ws.append([
            row.get("crew_log_id"),
            str(row.get("work_date")),
            row.get("site_id"),
            row.get("site_name"),
            row.get("vehicle_id"),
            row.get("vehicle_plate"),
            employees,
            float(row.get("work_hours") or 0.0),
            float(row.get("travel_hours") or 0.0),
            float(row.get("km") or 0.0),
            int(row.get("segments_count") or 0),
            round(float(row.get("km") or 0.0) * float(rate_per_km or 0.0), 2),
        ])
        total_work += float(row.get("work_hours") or 0.0)
        total_travel += float(row.get("travel_hours") or 0.0)
        total_km += float(row.get("km") or 0.0)
        total_cost += float(row.get("km") or 0.0) * float(rate_per_km or 0.0)

    ws.append([])
    ws.append(["SUMA", "", "", "", "", "", "", round(total_work, 2), round(total_travel, 2), round(total_km, 2), "", round(total_cost, 2)])

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 18

    from io import BytesIO
    from fastapi.responses import StreamingResponse
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"day_report_{date}.xlsx"
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




@router.get("/reports/range.xlsx")
def report_range_xlsx(
    date_from: date,
    date_to: date,
    db: Session = Depends(get_db),
):
    if date_to < date_from:
        raise HTTPException(status_code=422, detail="date_to must be >= date_from")

    from datetime import timedelta

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from starlette.responses import StreamingResponse

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    headers = [
        "date",
        "crew_log_id",
        "site_id",
        "site_name",
        "vehicle_id",
        "vehicle_plate",
        "employees",
        "work_hours",
        "travel_hours",
        "km",
        "segments_count",
        "travel_cost",
    ]
    ws_sum.append(headers)

    cur = date_from
    total_work = 0.0
    total_travel = 0.0
    total_km = 0.0
    total_cost = 0.0
    total_segments = 0

    while cur <= date_to:
        day = report_day(date=cur, db=db)
        crew_logs = day.get("crew_logs", [])

        ws_day = wb.create_sheet(title=str(cur))
        ws_day.append(headers[1:])

        day_work = 0.0
        day_travel = 0.0
        day_km = 0.0
        day_segments = 0

        for row in crew_logs:
            employees = ", ".join(row.get("employees") or [])

            w = float(row.get("work_hours") or 0.0)
            t = float(row.get("travel_hours") or 0.0)
            k = float(row.get("km") or 0.0)
            sc = int(row.get("segments_count") or 0)

            ws_sum.append([
                str(cur),
                row.get("crew_log_id"),
                row.get("site_id"),
                row.get("site_name"),
                row.get("vehicle_id"),
                row.get("vehicle_plate"),
                employees,
                w,
                t,
                k,
                sc,
            ])

            ws_day.append([
                row.get("crew_log_id"),
                row.get("site_id"),
                row.get("site_name"),
                row.get("vehicle_id"),
                row.get("vehicle_plate"),
                employees,
                w,
                t,
                k,
                sc,
            ])

            day_work += w
            day_travel += t
            day_km += k
            day_segments += sc

        ws_day.append([])
        ws_day.append([
            "SUMA",
            "",
            "",
            "",
            "",
            "",
            round(day_work, 2),
            round(day_travel, 2),
            round(day_km, 2),
            day_segments,
        ])

        for col in range(1, len(headers[1:]) + 1):
            ws_day.column_dimensions[get_column_letter(col)].width = 18

        total_work += day_work
        total_travel += day_travel
        total_km += day_km
        total_segments += day_segments

        cur = cur + timedelta(days=1)

    ws_sum.append([])
    ws_sum.append([
        "SUMA",
        "",
        "",
        "",
        "",
        "",
        "",
        round(total_work, 2),
        round(total_travel, 2),
        round(total_km, 2),
        total_segments,
    ])

    for col in range(1, len(headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"range_report_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




@router.get("/reports/employee", response_model=EmployeeReportOut)
def report_employee(
    employee_id: int,
    date_from: date,
    date_to: date,
    db: Session = Depends(get_db),
):
    if date_to < date_from:
        raise HTTPException(status_code=422, detail="date_to must be >= date_from")

    from datetime import timedelta

    emp = db.query(TkEmployee).filter(TkEmployee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    result_days = []

    cur = date_from
    total_work = 0
    total_travel = 0
    total_km = 0.0

    while cur <= date_to:
        segments = (
            db.query(TkCrewWorkSegment)
            .join(TkCrewLog, TkCrewLog.id == TkCrewWorkSegment.crew_log_id)
            .join(TkCrewLogMember, TkCrewLogMember.crew_log_id == TkCrewLog.id)
            .filter(TkCrewLogMember.employee_id == employee_id)
            .filter(TkCrewLog.work_date == cur)
            .filter(TkCrewWorkSegment.end_at.isnot(None))
            .all()
        )

        work_min = 0
        travel_min = 0
        km = 0.0

        for seg in segments:
            if not seg.start_at or not seg.end_at:
                continue

            minutes = int((seg.end_at - seg.start_at).total_seconds() // 60)
            if minutes < 0:
                minutes = 0
            minutes_15 = ((minutes + 14) // 15) * 15

            if str(getattr(seg, "segment_type", None) or "").lower() == "travel":
                travel_min += minutes_15
                km += float(seg.distance_km or 0.0)
            else:
                work_min += minutes_15

        if work_min or travel_min or km:
            result_days.append({
                "date": cur,
                "work_minutes": work_min,
                "work_hours": round(work_min / 60, 2),
                "travel_minutes": travel_min,
                "travel_hours": round(travel_min / 60, 2),
                "km": round(km, 2),
            })

            total_work += work_min
            total_travel += travel_min
            total_km += km

        cur = cur + timedelta(days=1)

    return {
        "employee_id": emp.id,
        "employee_name": emp.full_name,
        "date_from": date_from,
        "date_to": date_to,
        "total_work_hours": round(total_work / 60, 2),
        "total_travel_hours": round(total_travel / 60, 2),
        "total_km": round(total_km, 2),
        "days": result_days,
    }


@router.get("/reports/employee.xlsx")
def report_employee_xlsx(
    employee_id: int,
    date_from: date,
    date_to: date,
    rate_per_km: float = 0.0,
    db: Session = Depends(get_db),
):
    data = report_employee(
        employee_id=employee_id,
        date_from=date_from,
        date_to=date_to,
        db=db,
    )

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from starlette.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee"

    headers = [
        "date",
        "work_hours",
        "travel_hours",
        "km",
        "travel_cost",
    ]
    ws.append(headers)

    total_cost = 0.0

    for d in data.get("days", []):
        cost = round(float(d.get("km") or 0.0) * float(rate_per_km or 0.0), 2)
        total_cost += cost
        ws.append([
            str(d.get("date")),
            float(d.get("work_hours") or 0.0),
            float(d.get("travel_hours") or 0.0),
            float(d.get("km") or 0.0),
            cost,
        ])

    ws.append([])
    ws.append([
        "SUMA",
        data.get("total_work_hours"),
        data.get("total_travel_hours"),
        data.get("total_km"),
        round(total_cost, 2),
    ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"employee_{employee_id}_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=""{filename}""'},
    )











@router.get("/reports/day.pdf")
def report_day_pdf(date: date, db: Session = Depends(get_db)):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    data = report_day(date=date, db=db)
    rep_date = str(data.get("date", date))
    crew_logs = data.get("crew_logs", [])

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph(f"Raport dzienny – {rep_date}", styles["Title"]))
    elems.append(Spacer(1, 10))

    table_data = [[
        "Brygada",
        "Budowa",
        "Pojazd",
        "Praca [h]",
        "Dojazd [h]",
        "Km"
    ]]

    for c in crew_logs:
        table_data.append([
            str(c.get("crew_log_id","")),
            (c.get("site_name") or ""),
            (c.get("vehicle_plate") or ""),
            f"{float(c.get("work_hours") or 0.0):.2f}",
            f"{float(c.get("travel_hours") or 0.0):.2f}",
            f"{float(c.get("km") or 0.0):.1f}",
        ])

    tbl = Table(table_data, colWidths=[25*mm, 45*mm, 35*mm, 25*mm, 25*mm, 20*mm])
    tbl.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (3,1), (-1,-1), "RIGHT"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
        ("TOPPADDING", (0,0), (-1,0), 6),
    ]))

    elems.append(tbl)
    doc.build(elems, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)

    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=day_{rep_date}.pdf"},
    )




@router.get("/reports/range.pdf")
def report_range_pdf(date_from: date, date_to: date, rate_per_km: float = 0.0, db: Session = Depends(get_db)):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors

    data = report_range(date_from=date_from, date_to=date_to, db=db)

    df = str(data.get("date_from", date_from))
    dt = str(data.get("date_to", date_to))

    total_minutes = int(data.get("total_minutes") or 0)
    total_hours = float(data.get("total_hours") or 0.0)

    work_minutes = int(data.get("work_minutes") or 0)
    work_hours = float(data.get("work_hours") or 0.0)

    travel_minutes = int(data.get("travel_minutes") or 0)
    travel_hours = float(data.get("travel_hours") or 0.0)

    vehicles = data.get("vehicles") or []
    try:
        vehicles = sorted(vehicles, key=lambda x: float((x or {}).get("km") or 0.0), reverse=True)
    except Exception:
        pass

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=12*mm,
        bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph(f"Raport okresowy – {df} do {dt}", styles["Title"]))
    elems.append(Spacer(1, 8))

    sum_tbl = Table([
        ["Suma", "Min", "Godz"],
        ["Łącznie", str(total_minutes), f"{total_hours:.2f}"],
        ["Praca", str(work_minutes), f"{work_hours:.2f}"],
        ["Dojazd", str(travel_minutes), f"{travel_hours:.2f}"],
    ], colWidths=[35*mm, 35*mm, 35*mm])

    sum_tbl.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
        ("TOPPADDING", (0,0), (-1,0), 6),
    ]))

    elems.append(sum_tbl)
    elems.append(Spacer(1, 12))

    elems.append(Paragraph("Pojazdy (Top 20 po km)", styles["Heading2"]))
    elems.append(Spacer(1, 6))

    v_rows = [["Pojazd", "Rejestracja", "Km", "Godz", "Praca [h]", "Dojazd [h]", "Segmenty", "Koszt"]]

    top = vehicles[:20]
    for v in top:
        v = v or {}
        plate = str(v.get("plate") or "")
        veh_id = str(v.get("vehicle_id") or "")
        km = float(v.get("km") or 0.0)
        hours = float(v.get("hours") or 0.0)
        wh = float(v.get("work_hours") or 0.0)
        th = float(v.get("travel_hours") or 0.0)
        segs = int(v.get("segments") or 0)
        cost = km * float(rate_per_km or 0.0)
        cost_str = f"{cost:.2f}" if float(rate_per_km or 0.0) > 0 else ""
        v_rows.append([veh_id, plate, f"{km:.1f}", f"{hours:.2f}", f"{wh:.2f}", f"{th:.2f}", str(segs), cost_str])

    v_tbl = Table(v_rows, colWidths=[18*mm, 30*mm, 18*mm, 18*mm, 20*mm, 20*mm, 18*mm, 20*mm])
    v_tbl.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,0), 5),
        ("TOPPADDING", (0,0), (-1,0), 5),
    ]))

    elems.append(v_tbl)

    doc.build(elems, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    bio.seek(0)

    fn = f"range_{df}_to_{dt}.pdf".replace(":", "-")
    return StreamingResponse(
        bio,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={fn}"},
    )


def _pdf_footer(canvas, doc):
    from reportlab.lib.units import mm
    from datetime import datetime
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    left = 15 * mm
    right = doc.pagesize[0] - 15 * mm
    canvas.drawString(left, 10 * mm, "Eko Instal-Went | NIP 6343018528")
    canvas.drawRightString(right, 10 * mm, f"Wygenerowano: {ts} | Strona {canvas.getPageNumber()}")
    canvas.restoreState()




@router.get("/reports/payroll.xlsx")
def report_payroll_xlsx(date_from: date, date_to: date, db: Session = Depends(get_db)):
    from os import getenv
    if str(getenv("ENABLE_XLSX_EXPORT", "1")).lower() not in ("1","true","yes","y","t"):
        raise HTTPException(status_code=403, detail="XLSX export disabled")

    from io import BytesIO
    from fastapi.responses import StreamingResponse

    rows = _fetch_payroll_rows_sql(db, date_from, date_to)
    wb = _build_payroll_workbook(rows)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"payroll_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _fetch_payroll_rows_sql(db: Session, date_from: date, date_to: date):
    q = (
        db.query(
            TkCrewWorkSegment.id.label("segment_id"),
            TkCrewWorkSegment.crew_log_id.label("crew_log_id"),
            TkCrewWorkSegment.start_at.label("start_at"),
            TkCrewWorkSegment.end_at.label("end_at"),
            TkCrewWorkSegment.segment_type.label("segment_type"),
            TkCrewWorkSegment.distance_km.label("distance_km"),
            TkCrewWorkSegment.site_id.label("site_id"),
            TkCrewLog.work_date.label("work_date"),
            TkCrewLog.vehicle_id.label("vehicle_id"),
            TkCrewLogMember.employee_id.label("employee_id"),
            TkEmployee.full_name.label("employee_name"),
            TkVehicle.plate.label("vehicle_plate"),
            TkSite.name.label("site_name"),
        )
        .join(TkCrewLog, TkCrewLog.id == TkCrewWorkSegment.crew_log_id)
        .join(TkCrewLogMember, TkCrewLogMember.crew_log_id == TkCrewLog.id)
        .join(TkEmployee, TkEmployee.id == TkCrewLogMember.employee_id)
        .outerjoin(TkVehicle, TkVehicle.id == TkCrewLog.vehicle_id)
        .outerjoin(TkSite, TkSite.id == TkCrewWorkSegment.site_id)
        .filter(and_(TkCrewLog.work_date >= date_from, TkCrewLog.work_date <= date_to))
        .order_by(TkCrewLog.work_date.asc(), TkEmployee.full_name.asc(), TkCrewWorkSegment.start_at.asc())
    )
    return q.all()

def _build_payroll_workbook(rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws_segments = wb.active
    ws_segments.title = "Segments"
    ws_payroll = wb.create_sheet("Payroll")
    ws_totals = wb.create_sheet("Totals")
    ws_warn = wb.create_sheet("Warnings")

    seg_headers = [
        "work_date",
        "employee_id",
        "employee_name",
        "crew_log_id",
        "segment_id",
        "segment_type",
        "start_at",
        "end_at",
        "minutes_raw",
        "minutes_rounded_15",
        "hours_rounded",
        "km_travel",
        "vehicle_plate",
        "site_id",
        "site_name",
    ]
    ws_segments.append(seg_headers)

    payroll_headers = [
        "work_date",
        "employee_id",
        "employee_name",
        "work_minutes_rounded",
        "work_hours_rounded",
        "travel_minutes_rounded",
        "travel_hours_rounded",
        "km_travel",
    ]
    ws_payroll.append(payroll_headers)

    totals_headers = [
        "employee_id",
        "employee_name",
        "work_minutes_rounded",
        "work_hours_rounded",
        "travel_minutes_rounded",
        "travel_hours_rounded",
        "km_travel",
    ]
    ws_totals.append(totals_headers)

    warn_headers = [
        "level",
        "code",
        "work_date",
        "employee_id",
        "employee_name",
        "crew_log_id",
        "segment_id",
        "segment_type",
        "start_at",
        "end_at",
        "distance_km_db",
        "minutes_raw",
        "note",
    ]
    ws_warn.append(warn_headers)

    payroll_map = {}
    totals_map = {}

    def add_warn(level, code, r, seg_type, minutes_raw, note):
        ws_warn.append([
            level,
            code,
            getattr(r, "work_date", None),
            getattr(r, "employee_id", None),
            getattr(r, "employee_name", None),
            getattr(r, "crew_log_id", None),
            getattr(r, "segment_id", None),
            seg_type,
            getattr(r, "start_at", None),
            getattr(r, "end_at", None),
            float(getattr(r, "distance_km", 0) or 0),
            minutes_raw,
            note,
        ])

    for r in rows:
        seg_type = _norm_seg_type(getattr(r, "segment_type", None))
        if seg_type not in ("work", "travel"):
            add_warn("ERROR", "SEGMENT_TYPE_INVALID", r, seg_type, 0, "segment_type must be 'work' or 'travel'")
            seg_type = "work"

        start_at = getattr(r, "start_at", None)
        end_at = getattr(r, "end_at", None)

        if not start_at or not end_at:
            add_warn("ERROR", "MISSING_TIME", r, seg_type, 0, "start_at and end_at are required for payroll")
            minutes_raw = 0
        else:
            minutes_raw = _safe_minutes(start_at, end_at)
            if end_at < start_at:
                add_warn("ERROR", "NEGATIVE_DURATION", r, seg_type, minutes_raw, "end_at < start_at (clamped to 0)")

        minutes_rounded = _ceil_to_15(minutes_raw)
        hours_rounded = minutes_rounded / 60.0

        distance_db = float(getattr(r, "distance_km", 0) or 0)
        if distance_db < 0:
            add_warn("ERROR", "KM_NEGATIVE", r, seg_type, minutes_raw, "distance_km < 0 (forced to 0)")
            distance_db = 0.0

        if seg_type == "travel":
            km_travel = distance_db
            if km_travel == 0 and minutes_rounded > 0:
                add_warn("WARN", "TRAVEL_KM_ZERO", r, seg_type, minutes_raw, "travel segment has time but km=0")
        else:
            km_travel = 0.0
            if distance_db > 0:
                add_warn("ERROR", "WORK_HAS_KM", r, seg_type, minutes_raw, "work segment has km>0 (forced to 0)")

        if minutes_raw > 24 * 60:
            add_warn("WARN", "DURATION_GT_24H", r, seg_type, minutes_raw, "segment duration exceeds 24h")

        if getattr(r, "employee_id", None) is None:
            add_warn("ERROR", "MISSING_EMPLOYEE", r, seg_type, minutes_raw, "employee_id is null (join issue)")

        if getattr(r, "work_date", None) is None:
            add_warn("ERROR", "MISSING_WORK_DATE", r, seg_type, minutes_raw, "work_date is null (join issue)")

        if getattr(r, "crew_log_id", None) is None:
            add_warn("ERROR", "MISSING_CREW_LOG", r, seg_type, minutes_raw, "crew_log_id is null (join issue)")

        ws_segments.append([
            getattr(r, "work_date", None),
            getattr(r, "employee_id", None),
            getattr(r, "employee_name", None),
            getattr(r, "crew_log_id", None),
            getattr(r, "segment_id", None),
            seg_type,
            start_at,
            end_at,
            minutes_raw,
            minutes_rounded,
            hours_rounded,
            km_travel,
            getattr(r, "vehicle_plate", None),
            getattr(r, "site_id", None),
            getattr(r, "site_name", None),
        ])

        day_key = (getattr(r, "work_date", None), getattr(r, "employee_id", None), getattr(r, "employee_name", None))
        if day_key not in payroll_map:
            payroll_map[day_key] = {"work_min": 0, "travel_min": 0, "km": 0.0}

        if seg_type == "travel":
            payroll_map[day_key]["travel_min"] += minutes_rounded
            payroll_map[day_key]["km"] += km_travel
        else:
            payroll_map[day_key]["work_min"] += minutes_rounded

        emp_key = (getattr(r, "employee_id", None), getattr(r, "employee_name", None))
        if emp_key not in totals_map:
            totals_map[emp_key] = {"work_min": 0, "travel_min": 0, "km": 0.0}

        if seg_type == "travel":
            totals_map[emp_key]["travel_min"] += minutes_rounded
            totals_map[emp_key]["km"] += km_travel
        else:
            totals_map[emp_key]["work_min"] += minutes_rounded

    for (d, emp_id, emp_name), agg in sorted(payroll_map.items(), key=lambda x: (x[0][0], x[0][2] or "")):
        work_h = agg["work_min"] / 60.0
        travel_h = agg["travel_min"] / 60.0
        ws_payroll.append([d, emp_id, emp_name, agg["work_min"], work_h, agg["travel_min"], travel_h, agg["km"]])

    for (emp_id, emp_name), agg in sorted(totals_map.items(), key=lambda x: (x[0][1] or "")):
        work_h = agg["work_min"] / 60.0
        travel_h = agg["travel_min"] / 60.0
        ws_totals.append([emp_id, emp_name, agg["work_min"], work_h, agg["travel_min"], travel_h, agg["km"]])

    return wb





def _norm_seg_type(v):
    s = str(v or "").strip().lower()
    if s in ("travel", "drive", "driving", "jazda", "dojazd"):
        return "travel"
    return "work"

def _safe_minutes(start_at, end_at):
    if not start_at or not end_at:
        return 0
    try:
        delta = end_at - start_at
        secs = getattr(delta, "total_seconds", None)
        if callable(secs):
            s = secs()
        else:
            s = delta.total_seconds()
        if s is None:
            return 0
        m = int(s // 60)
        return m if m > 0 else 0
    except Exception:
        return 0

def _ceil_to_15(minutes: int) -> int:
    try:
        m = int(minutes or 0)
    except Exception:
        return 0
    if m <= 0:
        return 0
    return ((m + 14) // 15) * 15
